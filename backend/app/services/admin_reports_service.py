import re
from datetime import datetime, timezone
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
# pyrefly: ignore [missing-import]
from sqlalchemy.orm import Session

from app.models.evaluation_session import EvaluationSession
from app.schemas.admin_reports import AdminReportsSummaryRead, ReportSessionItemRead
from app.services.admin_dashboard_service import AdminDashboardService
from app.repositories.evaluation_session_repository import EvaluationSessionRepository


class AdminReportsService:
    FINISHED_STATUSES = {"completed", "expired"}

    def __init__(self, db: Session) -> None:
        self.db = db
        self.repository = EvaluationSessionRepository(db)
        self.dashboard_service = AdminDashboardService(db)

    def get_reports_summary(self) -> AdminReportsSummaryRead:
        finished_sessions = self._get_finished_sessions()
        return AdminReportsSummaryRead(
            generated_at=datetime.now(timezone.utc),
            evaluated_candidates_count=len({session.candidate_id for session in finished_sessions}),
            total_finished_sessions=len(finished_sessions),
            average_score_percentage=self._calculate_average_score_percentage(finished_sessions),
            average_time_seconds=self._calculate_average_time_seconds(finished_sessions),
            sessions=[self._build_report_session_item(session) for session in finished_sessions],
        )

    def build_general_excel(self) -> tuple[BytesIO, str]:
        dashboard = self.dashboard_service.get_dashboard_summary()
        finished_sessions = self._get_finished_sessions()

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Resumen"

        summary_rows = [
            ("Reporte", "General de evaluaciones"),
            ("Generado", self._format_datetime(datetime.now(timezone.utc))),
            ("Candidatos evaluados", dashboard.evaluated_candidates_count),
            ("Sesiones cerradas", dashboard.completed_sessions_count),
            ("Promedio general", f"{dashboard.average_score_percentage:.2f}%"),
            ("Tiempo promedio", self._format_duration(dashboard.average_time_seconds)),
            ("Mejor candidato", dashboard.best_candidate_name or "Sin datos"),
            (
                "Puntaje del mejor candidato",
                (
                    f"{dashboard.best_candidate_score_percentage:.2f}%"
                    if dashboard.best_candidate_score_percentage is not None
                    else "Sin datos"
                ),
            ),
        ]
        for key, value in summary_rows:
            summary_sheet.append([key, value])
        self._style_headerless_key_value_sheet(summary_sheet)

        sessions_sheet = workbook.create_sheet("Sesiones")
        sessions_sheet.append(
            [
                "ID sesion",
                "Candidato",
                "Plantilla",
                "Estado",
                "Puntaje",
                "Porcentaje",
                "Respondidas",
                "Omitidas",
                "Tiempo consumido",
                "Inicio",
                "Cierre",
                "Cierre por tiempo",
            ]
        )
        for session in finished_sessions:
            sessions_sheet.append(
                [
                    session.id,
                    self._build_candidate_name(session),
                    session.evaluation_template.name,
                    session.status,
                    session.total_score if session.total_score is not None else 0,
                    self._calculate_session_score_percentage(session),
                    session.answered_questions_count,
                    session.omitted_questions_count,
                    session.consumed_time_seconds,
                    self._format_datetime(session.started_at),
                    self._format_datetime(session.submitted_at),
                    "Si" if session.completed_by_timeout else "No",
                ]
            )
        self._style_table_sheet(sessions_sheet)

        categories_sheet = workbook.create_sheet("Categorias")
        categories_sheet.append(
            [
                "Categoria",
                "Promedio",
                "Tiempo promedio",
                "Reactivos evaluados",
                "Sesiones evaluadas",
            ]
        )
        for category in dashboard.category_averages:
            categories_sheet.append(
                [
                    category.category_name,
                    category.average_score_percentage,
                    category.average_time_seconds,
                    category.total_questions,
                    category.evaluated_sessions,
                ]
            )
        self._style_table_sheet(categories_sheet)

        ranking_sheet = workbook.create_sheet("Ranking")
        ranking_sheet.append(
            [
                "Posicion",
                "Candidato",
                "Correo",
                "Intentos",
                "Promedio",
                "Mejor puntaje",
                "Tiempo promedio",
                "Ultima plantilla",
                "Ultimo estado",
                "Ultimo cierre",
            ]
        )
        for index, item in enumerate(dashboard.ranking, start=1):
            ranking_sheet.append(
                [
                    index,
                    item.candidate_name,
                    item.email or "",
                    item.attempts_count,
                    item.average_score_percentage,
                    item.best_score_percentage,
                    item.average_time_seconds,
                    item.last_template_name or "",
                    item.last_status or "",
                    self._format_datetime(item.last_submitted_at),
                ]
            )
        self._style_table_sheet(ranking_sheet)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output, f"reporte_general_dsepc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    def build_general_pdf(self) -> tuple[BytesIO, str]:
        dashboard = self.dashboard_service.get_dashboard_summary()
        finished_sessions = self._get_finished_sessions()

        output = BytesIO()
        document = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements: list = []

        if "Center" not in styles:
            from reportlab.lib.styles import ParagraphStyle
            styles.add(ParagraphStyle(name='Center', alignment=1))
            
        banner_data = [[Paragraph("<font color='white' size='18'><b>REPORTE GENERAL DSEPC</b></font>", styles["Center"])]]
        banner = Table(banner_data, colWidths=[700])
        banner.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#0f172a")),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ]))
        elements.append(banner)
        elements.append(Spacer(1, 20))
        
        elements.append(
            Paragraph(
                f"Generado: {self._format_datetime(datetime.now(timezone.utc))}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 12))

        aptos_count = sum(1 for s in finished_sessions if self._calculate_session_score_percentage(s) >= 80.0)
        
        summary_table = Table(
            [
                ["Indicador", "Valor"],
                ["Candidatos evaluados", dashboard.evaluated_candidates_count],
                ["Sesiones cerradas", dashboard.completed_sessions_count],
                ["Candidatos Aptos", f"{aptos_count} (Umbral: 80%)"],
                ["Promedio general", f"{dashboard.average_score_percentage:.2f}%"],
                ["Tiempo promedio", self._format_duration(dashboard.average_time_seconds)],
                ["Mejor candidato", dashboard.best_candidate_name or "Sin datos"],
            ],
            colWidths=[200, 260],
        )
        self._apply_pdf_table_style(summary_table)
        elements.append(summary_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("Promedio por categoria", styles["Heading2"]))
        category_table = Table(
            [
                [
                    "Categoria",
                    "Promedio",
                    "Tiempo promedio",
                    "Reactivos",
                    "Sesiones",
                ]
            ]
            + [
                [
                    item.category_name,
                    f"{item.average_score_percentage:.2f}%",
                    self._format_duration(item.average_time_seconds),
                    item.total_questions,
                    item.evaluated_sessions,
                ]
                for item in dashboard.category_averages
            ],
            repeatRows=1,
        )
        self._apply_pdf_table_style(category_table)
        elements.append(category_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("Top 10 ranking", styles["Heading2"]))
        ranking_table = Table(
            [
                ["Pos.", "Candidato", "Promedio", "Mejor", "Estatus", "Intentos", "Tiempo prom."]
            ]
            + [
                [
                    index,
                    item.candidate_name,
                    f"{item.average_score_percentage:.2f}%",
                    f"{item.best_score_percentage:.2f}%",
                    "APTO" if item.best_score_percentage >= 80.0 else "NO APTO",
                    item.attempts_count,
                    self._format_duration(item.average_time_seconds),
                ]
                for index, item in enumerate(dashboard.ranking[:10], start=1)
            ],
            repeatRows=1,
            colWidths=[35, 200, 75, 75, 75, 60, 90]
        )
        self._apply_pdf_table_style(ranking_table)
        elements.append(ranking_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("Ultimas sesiones cerradas", styles["Heading2"]))
        sessions_table = Table(
            [
                ["Sesion", "Candidato", "Plantilla", "Estatus", "Porcentaje", "Tiempo", "Cierre"]
            ]
            + [
                [
                    session.id,
                    self._build_candidate_name(session),
                    session.evaluation_template.name,
                    "APTO" if self._calculate_session_score_percentage(session) >= 80.0 else "NO APTO",
                    f"{self._calculate_session_score_percentage(session):.2f}%",
                    self._format_duration(session.consumed_time_seconds),
                    self._format_datetime(session.submitted_at),
                ]
                for session in finished_sessions[:20]
            ],
            repeatRows=1,
            colWidths=[40, 160, 150, 70, 75, 60, 130]
        )
        self._apply_pdf_table_style(sessions_table)
        elements.append(sessions_table)

        document.build(elements)
        output.seek(0)
        return output, f"reporte_general_dsepc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    def build_session_excel(self, session_id: int) -> tuple[BytesIO, str]:
        session = self._get_session_or_404(session_id)
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Resumen"

        summary_rows = [
            ("Sesion", session.id),
            ("Candidato", self._build_candidate_name(session)),
            ("Plantilla", session.evaluation_template.name),
            ("Estado", session.status),
            ("Puntaje", session.total_score if session.total_score is not None else 0),
            ("Porcentaje", f"{self._calculate_session_score_percentage(session):.2f}%"),
            ("Respondidas", session.answered_questions_count),
            ("Omitidas", session.omitted_questions_count),
            ("Tiempo consumido", self._format_duration(session.consumed_time_seconds)),
            ("Inicio", self._format_datetime(session.started_at)),
            ("Cierre", self._format_datetime(session.submitted_at)),
            ("Cierre por tiempo", "Si" if session.completed_by_timeout else "No"),
        ]
        for key, value in summary_rows:
            summary_sheet.append([key, value])
        self._style_headerless_key_value_sheet(summary_sheet)

        categories_sheet = workbook.create_sheet("Categorias")
        categories_sheet.append(
            [
                "Categoria",
                "Reactivos",
                "Respondidas",
                "Omitidas",
                "Correctas",
                "Incorrectas",
                "Puntaje obtenido",
                "Puntaje posible",
            ]
        )
        for item in self._build_session_category_rows(session):
            categories_sheet.append(item)
        self._style_table_sheet(categories_sheet)

        questions_sheet = workbook.create_sheet("Preguntas")
        questions_sheet.append(
            [
                "Orden seccion",
                "Seccion",
                "Orden pregunta",
                "Categoria",
                "Pregunta",
                "Respuesta seleccionada",
                "Respuesta correcta",
                "Respondida",
                "Correcta",
                "Omitida",
                "Tiempo",
                "Puntaje",
            ]
        )
        for section in session.sections:
            for session_question in section.questions:
                questions_sheet.append(
                    [
                        section.sort_order,
                        section.title,
                        session_question.sort_order,
                        session_question.question.category.name,
                        session_question.question.statement,
                        session_question.selected_answer or "",
                        session_question.question.correct_answer,
                        "Si" if session_question.is_answered else "No",
                        (
                            "Si"
                            if session_question.is_correct
                            else "No" if session_question.is_correct is False else ""
                        ),
                        "Si" if session_question.was_omitted else "No",
                        session_question.time_spent_seconds,
                        session_question.question.score,
                    ]
                )
        self._style_table_sheet(questions_sheet)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = (
            f"reporte_individual_{self._safe_filename(self._build_candidate_name(session))}"
            f"_sesion_{session.id}.xlsx"
        )
        return output, filename

    def build_session_pdf(self, session_id: int) -> tuple[BytesIO, str]:
        session = self._get_session_or_404(session_id)
        output = BytesIO()
        document = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        if "Center" not in styles:
            from reportlab.lib.styles import ParagraphStyle
            styles.add(ParagraphStyle(name='Center', alignment=1))
            
        elements: list = []

        banner_data = [[Paragraph("<font color='white' size='18'><b>REPORTE INDIVIDUAL DSEPC</b></font>", styles["Center"])]]
        banner = Table(banner_data, colWidths=[700])
        banner.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#0f172a")),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ]))
        elements.append(banner)
        elements.append(Spacer(1, 20))

        info_data = [
            [f"Candidato: {self._build_candidate_name(session)}"],
            [f"Plantilla: {session.evaluation_template.name}"]
        ]
        info_table = Table(info_data, colWidths=[700], hAlign='LEFT')
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 11),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 15))

        score_percentage = self._calculate_session_score_percentage(session)
        is_apto = score_percentage >= 80.0
        badge_color = colors.HexColor("#16a34a") if is_apto else colors.HexColor("#dc2626")
        badge_text = "APTO" if is_apto else "NO APTO"

        badge_data = [[Paragraph(f"<font color='white' size='14'><b>DICTAMEN: {badge_text} ({score_percentage:.2f}%)</b></font>", styles["Center"])]]
        badge = Table(badge_data, colWidths=[300], hAlign='LEFT')
        badge.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), badge_color),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        elements.append(badge)
        elements.append(Spacer(1, 15))

        summary_table = Table(
            [
                ["Indicador", "Valor"],
                ["Sesion", session.id],
                ["Estado", session.status],
                ["Puntaje", session.total_score if session.total_score is not None else 0],
                ["Respondidas", session.answered_questions_count],
                ["Omitidas", session.omitted_questions_count],
                ["Tiempo consumido", self._format_duration(session.consumed_time_seconds)],
                ["Inicio", self._format_datetime(session.started_at)],
                ["Cierre", self._format_datetime(session.submitted_at)],
            ],
            colWidths=[170, 300],
        )
        self._apply_pdf_table_style(summary_table)
        elements.append(summary_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("Resultados por categoria", styles["Heading2"]))
        category_table = Table(
            [
                [
                    "Categoria",
                    "Reactivos",
                    "Correctas",
                    "Incorrectas",
                    "Omitidas",
                    "Puntaje",
                ]
            ]
            + [
                [
                    row[0],
                    row[1],
                    row[4],
                    row[5],
                    row[3],
                    f"{row[6]} / {row[7]}",
                ]
                for row in self._build_session_category_rows(session)
            ],
            repeatRows=1,
        )
        self._apply_pdf_table_style(category_table)
        elements.append(category_table)
        elements.append(PageBreak())
        elements.append(Paragraph("Detalle de preguntas", styles["Heading2"]))


        question_rows = [
            [
                "Sec.",
                "Pregunta",
                "Seleccionada",
                "Correcta",
                "Resultado",
                "Tiempo",
            ]
        ]
        for section in session.sections:
            for session_question in section.questions:
                question_rows.append(
                    [
                        str(section.sort_order),
                        Paragraph(
                            (session_question.question.statement or "")
                            .replace("&", "&amp;")
                            .replace("<", "&lt;")
                            .replace(">", "&gt;"),
                            styles["Normal"],
                        ),
                        Paragraph(
                            (session_question.selected_answer or "-")
                            .replace("&", "&amp;")
                            .replace("<", "&lt;")
                            .replace(">", "&gt;"),
                            styles["Normal"],
                        ),
                        Paragraph(
                            (session_question.question.correct_answer or "")
                            .replace("&", "&amp;")
                            .replace("<", "&lt;")
                            .replace(">", "&gt;"),
                            styles["Normal"],
                        ),
                        self._get_question_result_label(session_question, styles),
                        self._format_duration(session_question.time_spent_seconds),
                    ]
                )

        questions_table = Table(question_rows, repeatRows=1, colWidths=[35, 195, 95, 95, 205, 55])
        self._apply_pdf_table_style(questions_table)
        elements.append(questions_table)

        document.build(elements)
        output.seek(0)
        filename = (
            f"reporte_individual_{self._safe_filename(self._build_candidate_name(session))}"
            f"_sesion_{session.id}.pdf"
        )
        return output, filename

    def _get_question_result_label(self, session_question, styles) -> object:
        if session_question.question.question_type == "excel_practical" and session_question.practical_feedback:
            import json
            try:
                fb = json.loads(session_question.practical_feedback)
                lines = [f"<b>{fb.get('correct_cells', 0)}/{fb.get('total_cells', 0)} aciertos</b>"]
                
                criteria_results = fb.get("criteria_results", {})
                if criteria_results:
                    for c_name, c_data in criteria_results.items():
                        c_correct = c_data.get("correct", 0)
                        c_total = c_data.get("total", 0)
                        pct = int(c_data.get("success_rate", 0) * 100)
                        lines.append(f"• {c_name}: {c_correct}/{c_total} ({pct}%)")
                        
                from reportlab.platypus import Paragraph
                return Paragraph("<br/>".join(lines), styles["Normal"])
            except Exception:
                pass
        
        if session_question.is_correct:
            return "Correcta"
        if session_question.is_correct is False:
            return "Incorrecta"
        return "Omitida"

    def _get_finished_sessions(self) -> list[EvaluationSession]:
        sessions = self.repository.list_all_for_dashboard()
        return [session for session in sessions if session.status in self.FINISHED_STATUSES]

    def _get_session_or_404(self, session_id: int) -> EvaluationSession:
        session = self.repository.get_by_id(session_id)
        if not session:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="La sesion no existe.",
            )
        return session

    def _build_report_session_item(self, session: EvaluationSession) -> ReportSessionItemRead:
        return ReportSessionItemRead(
            session_id=session.id,
            candidate_name=self._build_candidate_name(session),
            template_name=session.evaluation_template.name,
            status=session.status,
            total_score=session.total_score,
            score_percentage=self._calculate_session_score_percentage(session),
            answered_questions=session.answered_questions_count,
            omitted_questions=session.omitted_questions_count,
            consumed_time_seconds=session.consumed_time_seconds,
            started_at=self._normalize_datetime(session.started_at),
            submitted_at=(
                self._normalize_datetime(session.submitted_at) if session.submitted_at else None
            ),
            completed_by_timeout=session.completed_by_timeout,
        )

    def _build_candidate_name(self, session: EvaluationSession) -> str:
        return f"{session.candidate.first_name} {session.candidate.last_name}".strip()

    def _calculate_average_score_percentage(self, sessions: list[EvaluationSession]) -> float:
        if not sessions:
            return 0.0
        return round(
            sum(self._calculate_session_score_percentage(session) for session in sessions)
            / len(sessions),
            2,
        )

    def _calculate_average_time_seconds(self, sessions: list[EvaluationSession]) -> float:
        if not sessions:
            return 0.0
        return round(
            sum(session.consumed_time_seconds for session in sessions) / len(sessions),
            2,
        )

    def _calculate_session_score_percentage(self, session: EvaluationSession) -> float:
        score_possible = 0.0
        score_obtained = 0.0

        for section in session.sections:
            for session_question in section.questions:
                score_possible += session_question.question.score
                if session_question.question.question_type == "excel_practical" and session_question.practical_feedback:
                    import json
                    try:
                        fb = json.loads(session_question.practical_feedback)
                        score_obtained += session_question.question.score * fb.get("success_rate", 0.0)
                    except Exception:
                        if session_question.is_correct:
                            score_obtained += session_question.question.score
                elif session_question.is_correct:
                    score_obtained += session_question.question.score

        if score_possible == 0:
            return 0.0

        return round((score_obtained / score_possible) * 100, 2)

    def _build_session_category_rows(self, session: EvaluationSession) -> list[list[object]]:
        category_metrics: dict[str, dict[str, float | int]] = {}

        for section in session.sections:
            for session_question in section.questions:
                category_name = session_question.question.category.name
                metrics = category_metrics.setdefault(
                    category_name,
                    {
                        "total_questions": 0,
                        "answered_questions": 0,
                        "omitted_questions": 0,
                        "correct_questions": 0,
                        "incorrect_questions": 0,
                        "score_obtained": 0.0,
                        "score_possible": 0.0,
                    },
                )

                metrics["total_questions"] += 1
                metrics["score_possible"] += session_question.question.score

                if session_question.is_answered:
                    metrics["answered_questions"] += 1
                    if session_question.is_correct:
                        metrics["correct_questions"] += 1
                        metrics["score_obtained"] += session_question.question.score
                    else:
                        metrics["incorrect_questions"] += 1
                elif session_question.was_omitted:
                    metrics["omitted_questions"] += 1

        return [
            [
                category_name,
                int(metrics["total_questions"]),
                int(metrics["answered_questions"]),
                int(metrics["omitted_questions"]),
                int(metrics["correct_questions"]),
                int(metrics["incorrect_questions"]),
                round(float(metrics["score_obtained"]), 2),
                round(float(metrics["score_possible"]), 2),
            ]
            for category_name, metrics in sorted(category_metrics.items())
        ]

    def _style_table_sheet(self, worksheet) -> None:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    def _style_headerless_key_value_sheet(self, worksheet) -> None:
        for cell in worksheet["A"]:
            cell.font = Font(bold=True)

        worksheet.column_dimensions["A"].width = 28
        worksheet.column_dimensions["B"].width = 40

    def _apply_pdf_table_style(self, table: Table) -> None:
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )

    def _format_datetime(self, value: datetime | None) -> str:
        if value is None:
            return "Sin dato"
        normalized = self._normalize_datetime(value)
        return normalized.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    def _format_duration(self, seconds: float | int) -> str:
        total_seconds = max(0, int(round(seconds)))
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        remaining_seconds = total_seconds % 60

        if hours:
            return f"{hours}h {minutes:02d}m"
        if minutes:
            return f"{minutes}m {remaining_seconds:02d}s"
        return f"{remaining_seconds}s"

    def _truncate_text(self, value: str, length: int) -> str:
        if len(value) <= length:
            return value
        return f"{value[: length - 3]}..."

    def _safe_filename(self, value: str) -> str:
        normalized = re.sub(r"[^A-Za-z0-9]+", "_", value.strip())
        return normalized.strip("_").lower() or "reporte"

    def _normalize_datetime(self, value: datetime) -> datetime:
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
