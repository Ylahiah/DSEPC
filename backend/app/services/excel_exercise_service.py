import json
from io import BytesIO
from pathlib import Path
from uuid import uuid4
import math

import pandas as pd
from openpyxl import load_workbook
from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.excel_exercise import ExcelExercise
from app.schemas.excel_exercise import (
    ExcelExerciseForm,
    ExcelExerciseRead,
)


class ExcelExerciseService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.settings = get_settings()

    def list_exercises(self) -> list[ExcelExerciseRead]:
        statement = select(ExcelExercise).order_by(ExcelExercise.created_at.desc())
        exercises = list(self.db.scalars(statement).all())
        return [self._build_read(exercise) for exercise in exercises]

    def create_exercise(
        self,
        payload: ExcelExerciseForm,
        workbook: UploadFile,
        solution_workbook: UploadFile,
    ) -> ExcelExerciseRead:
        self._ensure_unique_name(payload.name)
        workbook_bytes = self._read_uploaded_workbook(workbook)
        solution_bytes = self._read_uploaded_workbook(solution_workbook)
        
        expected_summary = self._analyze_solution_workbook(
            solution_bytes=solution_bytes,
            task_sheet_name=payload.task_sheet_name,
        )
        
        stored_path = self._store_workbook_bytes(workbook.filename or "ejercicio.xlsx", workbook_bytes)
        solution_stored_path = self._store_workbook_bytes(solution_workbook.filename or "solucion.xlsx", solution_bytes)

        exercise = ExcelExercise(
            name=payload.name,
            description=payload.description,
            instructions=payload.instructions,
            workbook_filename=workbook.filename or "ejercicio.xlsx",
            workbook_storage_path=str(stored_path),
            solution_filename=solution_workbook.filename or "solucion.xlsx",
            solution_storage_path=str(solution_stored_path),
            source_sheet_name=payload.source_sheet_name,
            task_sheet_name=payload.task_sheet_name,
            expected_summary_json=json.dumps(expected_summary, ensure_ascii=False),
            is_active=payload.is_active,
        )
        self.db.add(exercise)
        self.db.commit()
        self.db.refresh(exercise)
        return self._build_read(exercise)

    def update_exercise(
        self,
        exercise_id: int,
        payload: ExcelExerciseForm,
        workbook: UploadFile | None,
        solution_workbook: UploadFile | None,
    ) -> ExcelExerciseRead:
        exercise = self._get_or_404(exercise_id)
        self._ensure_unique_name(payload.name, excluded_id=exercise_id)

        exercise.name = payload.name
        exercise.description = payload.description
        exercise.instructions = payload.instructions
        exercise.source_sheet_name = payload.source_sheet_name
        exercise.task_sheet_name = payload.task_sheet_name
        exercise.is_active = payload.is_active

        if workbook is not None:
            workbook_bytes = self._read_uploaded_workbook(workbook)
            old_path = Path(exercise.workbook_storage_path)
            stored_path = self._store_workbook_bytes(
                workbook.filename or exercise.workbook_filename,
                workbook_bytes,
            )
            exercise.workbook_filename = workbook.filename or exercise.workbook_filename
            exercise.workbook_storage_path = str(stored_path)
            if old_path.exists():
                old_path.unlink()
        else:
            current_file = Path(exercise.workbook_storage_path)
            if not current_file.exists():
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="El archivo fuente del ejercicio ya no existe en el servidor.",
                )

        if solution_workbook is not None:
            solution_bytes = self._read_uploaded_workbook(solution_workbook)
            expected_summary = self._analyze_solution_workbook(
                solution_bytes=solution_bytes,
                task_sheet_name=payload.task_sheet_name,
            )
            old_sol_path = Path(exercise.solution_storage_path) if exercise.solution_storage_path else None
            solution_stored_path = self._store_workbook_bytes(
                solution_workbook.filename or "solucion.xlsx",
                solution_bytes,
            )
            exercise.solution_filename = solution_workbook.filename or "solucion.xlsx"
            exercise.solution_storage_path = str(solution_stored_path)
            exercise.expected_summary_json = json.dumps(expected_summary, ensure_ascii=False)
            if old_sol_path and old_sol_path.exists():
                old_sol_path.unlink()
        else:
            if not exercise.solution_storage_path:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Este es un ejercicio antiguo. Debes subir un Archivo Solucion para poder actualizarlo.",
                )
            current_sol_file = Path(exercise.solution_storage_path)
            if not current_sol_file.exists():
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="El archivo solucion del ejercicio ya no existe en el servidor.",
                )

        self.db.commit()
        self.db.refresh(exercise)
        return self._build_read(exercise)

    def delete_exercise(self, exercise_id: int) -> None:
        exercise = self._get_or_404(exercise_id)
        stored_path = Path(exercise.workbook_storage_path)
        sol_stored_path = Path(exercise.solution_storage_path) if exercise.solution_storage_path else None
        self.db.delete(exercise)
        self.db.commit()
        if stored_path.exists():
            stored_path.unlink()
        if sol_stored_path and sol_stored_path.exists():
            sol_stored_path.unlink()

    def set_status(self, exercise_id: int, is_active: bool) -> ExcelExerciseRead:
        exercise = self._get_or_404(exercise_id)
        exercise.is_active = is_active
        self.db.commit()
        self.db.refresh(exercise)
        return self._build_read(exercise)

    def get_download(self, exercise_id: int) -> tuple[BytesIO, str]:
        exercise = self._get_or_404(exercise_id)
        return self.get_download_for_exercise(exercise)

    def get_download_for_exercise(self, exercise: ExcelExercise) -> tuple[BytesIO, str]:
        stored_path = Path(exercise.workbook_storage_path)
        if not stored_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="El archivo del ejercicio no existe en el almacenamiento.",
            )

        return BytesIO(stored_path.read_bytes()), exercise.workbook_filename

    def validate_submission(
        self,
        exercise: ExcelExercise,
        workbook_bytes: bytes,
    ) -> tuple[bool, str]:
        original_path = Path(exercise.workbook_storage_path)
        if not original_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="El archivo original del ejercicio ya no existe en el servidor.",
            )

        original_bytes = original_path.read_bytes()
        self._ensure_source_sheet_integrity(
            original_bytes=original_bytes,
            submitted_bytes=workbook_bytes,
            source_sheet_name=exercise.source_sheet_name,
        )

        expected_summary = json.loads(exercise.expected_summary_json)
        self._ensure_task_results_match(
            workbook_bytes=workbook_bytes,
            task_sheet_name=exercise.task_sheet_name,
            expected_summary=expected_summary,
        )

        return (
            True,
            "Archivo validado correctamente. Los resultados coinciden con lo esperado.",
        )

    def _get_or_404(self, exercise_id: int) -> ExcelExercise:
        exercise = self.db.get(ExcelExercise, exercise_id)
        if not exercise:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="El ejercicio practico no existe.",
            )
        return exercise

    def _ensure_unique_name(self, name: str, excluded_id: int | None = None) -> None:
        statement = select(ExcelExercise).where(ExcelExercise.name == name)
        existing = self.db.scalar(statement)
        if existing and existing.id != excluded_id:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Ya existe un ejercicio practico con ese nombre.",
            )

    def _read_uploaded_workbook(self, workbook: UploadFile) -> bytes:
        filename = workbook.filename or ""
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El archivo del ejercicio debe estar en formato Excel (.xlsx o .xlsm).",
            )

        workbook_bytes = workbook.file.read()
        if not workbook_bytes:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El archivo del ejercicio esta vacio.",
            )
        return workbook_bytes

    def _analyze_solution_workbook(
        self,
        *,
        solution_bytes: bytes,
        task_sheet_name: str,
    ) -> dict[str, object]:
        try:
            wb = load_workbook(filename=BytesIO(solution_bytes), data_only=True)
        except Exception as error:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No fue posible abrir el archivo Excel de solucion.",
            ) from error

        if task_sheet_name not in wb.sheetnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"La hoja de solucion '{task_sheet_name}' no existe en el archivo.",
            )

        ws = wb[task_sheet_name]
        target_cells = {}
        import datetime
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    val = cell.value
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        val = val.isoformat()
                    target_cells[cell.coordinate] = val

        if not target_cells:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La hoja de solucion esta vacia. Debe contener los resultados esperados.",
            )

        return target_cells

    def _store_workbook_bytes(self, filename: str, workbook_bytes: bytes) -> Path:
        storage_dir = self.settings.excel_exercise_storage_dir
        storage_dir.mkdir(parents=True, exist_ok=True)
        suffix = Path(filename).suffix.lower() or ".xlsx"
        stored_path = storage_dir / f"{uuid4().hex}{suffix}"
        stored_path.write_bytes(workbook_bytes)
        return stored_path

    def _ensure_source_sheet_integrity(
        self,
        *,
        original_bytes: bytes,
        submitted_bytes: bytes,
        source_sheet_name: str,
    ) -> None:
        original_df = pd.read_excel(BytesIO(original_bytes), sheet_name=source_sheet_name).fillna("")
        submitted_df = pd.read_excel(BytesIO(submitted_bytes), sheet_name=source_sheet_name).fillna("")
        if not original_df.equals(submitted_df):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La hoja base del ejercicio fue modificada y debe conservarse sin cambios.",
            )

    def _ensure_task_results_match(
        self,
        *,
        workbook_bytes: bytes,
        task_sheet_name: str,
        expected_summary: dict[str, object],
    ) -> None:
        try:
            wb = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
        except Exception as error:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No fue posible leer el archivo Excel enviado.",
            ) from error

        if task_sheet_name not in wb.sheetnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"El archivo no contiene la hoja requerida: '{task_sheet_name}'.",
            )
            
        ws = wb[task_sheet_name]
        errors = []
        
        import datetime
        for coord, expected_val in expected_summary.items():
            cell = ws[coord]
            actual_val = cell.value
            if isinstance(actual_val, (datetime.datetime, datetime.date)):
                actual_val = actual_val.isoformat()
            
            if isinstance(expected_val, (int, float)) and isinstance(actual_val, (int, float)):
                if not math.isclose(expected_val, actual_val, rel_tol=1e-5):
                    errors.append(coord)
            else:
                if str(expected_val).strip() != str(actual_val).strip():
                    errors.append(coord)
                    
        if errors:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Los resultados no coinciden en {len(errors)} celdas. Ejemplo: celda {errors[0]}.",
            )

    def _build_read(self, exercise: ExcelExercise) -> ExcelExerciseRead:
        target_cells = json.loads(exercise.expected_summary_json)
        return ExcelExerciseRead(
            id=exercise.id,
            name=exercise.name,
            description=exercise.description,
            instructions=exercise.instructions,
            workbook_filename=exercise.workbook_filename,
            solution_filename=exercise.solution_filename,
            source_sheet_name=exercise.source_sheet_name,
            task_sheet_name=exercise.task_sheet_name,
            target_cells_count=len(target_cells),
            is_active=exercise.is_active,
            created_at=exercise.created_at,
            updated_at=exercise.updated_at,
        )
